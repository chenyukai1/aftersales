# -*- coding: utf-8 -*-
"""主数据导入工具：Excel → Frappe DocType（当前：配件价格表 Spare Part）。

配件价格表是登记自动带出 / M1 出库物料 / M2 追回判定 / M3 索赔清单的核心主档。

用法：
1. 用本模块生成的模板填写（本地 data/配件价格表导入模板.xlsx，勿改表头/顺序）。
2. 容器内执行：
   bench --site dev.localhost execute \\
     "frappe.get_attr('aftersales.after_sales.import_data.import_spare_parts')('/home/frappe/frappe-bench/apps/aftersales/data/配件价格表.xlsx')"
3. 返回统计 {created, updated, skipped, errors}。

特点：
- 幂等：按 K3编码 upsert（存在则更新其余字段，不重复建）。
- 供应商不存在时自动创建（All Supplier Groups）。
- 索赔需求仅接受「售后选项」配置内的合法值，非法值该行仅跳过索赔需求并提示。
- ERP物料编码：可选；提供且 Item 存在则绑定 erp_item（M1 出库用），缺失不影响导入。
"""
import frappe
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# 合法索赔需求（须与「售后选项」中「索赔需求」类别一致）
VALID_CLAIM_REQUIREMENTS = [
    "需提供旧件",
    "仅需清单",
    "供应商预赔无需清单&资料",
    "每月提供售后清单",
    "需资料",
]

# 模板列（顺序固定，勿改动）
TEMPLATE_COLS = [
    "K3编码", "E10品号", "配件品名", "供应商", "索赔需求", "ERP物料编码", "备注",
]
REQUIRED_COLS = {"K3编码", "配件品名"}
OPTIONAL_ITEM = "ERP物料编码"

# 列名别名（兼容真实 K3 导出的不同表头）
COL_ALIASES = {
    "k3_code": ["k3编码", "k3code", "k3", "编码", "物料编码", "配件编码", "旧件编码", "零件号"],
    "e10_code": ["e10品号", "e10code", "e10", "品号"],
    "part_name": ["配件品名", "品名", "配件名称", "名称", "零件名称", "物料名称"],
    "supplier": ["供应商", "供应商名称"],
    "claim_requirement": ["索赔需求"],
    "erp_item": ["erp物料编码", "erp物料", "erp编码", "物料代码", "erpitem"],
}

SAMPLE_ROWS = [
    # (K3编码, E10品号, 配件品名, 供应商, 索赔需求, ERP物料编码)
    ("31101166", "31.101.0166", "PU双轮 80*70mm 带轴承 (REACH认证)", "江苏威博", "供应商预赔无需清单&资料", "31101166"),
    ("31101130", "31.101.0130", "液压站总成 (VIBO) 2.0T", "江苏威博", "每月提供售后清单", "31101130"),
    ("31501014", "31.501.0014", "00842 油缸总成 (活塞杆直径Ø35)", "华昌液压", "需资料", "31501014"),
]


def make_template(path):
    """生成导入模板 xlsx（含下拉、示例、填写说明）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "配件价格表"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    sample_fill = PatternFill("solid", fgColor="FFF2CC")
    ws.append(TEMPLATE_COLS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in SAMPLE_ROWS:
        ws.append(list(row) + [""])
    for r in range(2, 2 + len(SAMPLE_ROWS)):
        for cell in ws[r]:
            cell.fill = sample_fill
    # 索赔需求下拉（第5列 E）
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(VALID_CLAIM_REQUIREMENTS) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="非法索赔需求",
        error="仅接受：" + " / ".join(VALID_CLAIM_REQUIREMENTS),
    )
    ws.add_data_validation(dv)
    dv.add(f"E2:E2000")
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFG", (14, 14, 40, 22, 24, 16, 20)):
        ws.column_dimensions[col].width = w

    # 填写说明
    doc = wb.create_sheet("填写说明")
    lines = [
        ("配件价格表导入说明", True),
        ("", False),
        ("1. 必填列：K3编码、配件品名；建议列：供应商、索赔需求（决定登记自动带出与索赔清单逻辑）。", False),
        ("2. K3编码即「老配件编码/新配件编码」登记时输入的编码，请与 K3 系统保持一致。", False),
        ("3. 索赔需求下拉可选：" + " / ".join(VALID_CLAIM_REQUIREMENTS), False),
        ("4. 供应商会自动创建（All Supplier Groups 分组）；若已在主档则直接关联。", False),
        ("5. ERP物料编码（可选）：填 ERP/Item 中已有的物料编码则绑定出库物料（M1 自动出库用）；", False),
        ("   留空不影响本次导入，可后续在「配件主档」中补充。", False),
        ("6. 黄色行=示例，导入时按 K3编码更新，可删除后粘贴真实数据。", False),
        ("7. 幂等：同一 K3编码 重复导入只会更新，不会重复建。", False),
        ("8. 填好后保存为 data/配件价格表.xlsx 并执行导入命令（见 README 或项目日志）。", False),
        ("", False),
        ("导入命令：", True),
        ("bench --site dev.localhost execute \"frappe.get_attr('aftersales.after_sales.import_data.import_spare_parts')('/home/frappe/frappe-bench/apps/aftersales/data/配件价格表.xlsx')\"", False),
    ]
    for text, bold in lines:
        doc.append([text])
        doc.cell(row=doc.max_row, column=1).font = Font(bold=bold)
    doc.column_dimensions["A"].width = 110
    wb.save(path)
    return path


def import_spare_parts(path):
    """批量导入配件价格表（幂等）。返回 {created, updated, skipped, errors:[...]}。"""
    stats = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        return {"error": f"无法打开 Excel：{e}", **stats}
    ws = wb["配件价格表"] if "配件价格表" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"error": "Excel 无数据行", **stats}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    # 表头 → 内部字段映射（大小写不敏感）
    col_idx = {}
    headers_lower = [h.lower() for h in headers]
    for field, aliases in COL_ALIASES.items():
        aliases_lower = {a.lower() for a in aliases}
        for i, h in enumerate(headers_lower):
            if h in aliases_lower:
                col_idx[field] = i
                break
    if "k3_code" not in col_idx:
        return {"error": f"未找到编码列，表头：{headers}", **stats}
    if "part_name" not in col_idx:
        return {"error": f"未找到配件品名列，表头：{headers}", **stats}
    supp_cache, created_suppliers = {}, []
    item_group = "Products"
    for r in rows[1:]:
        if not r or all(c is None or str(c).strip() == "" for c in r):
            continue
        k3 = str(r[col_idx["k3_code"]]).strip() if col_idx.get("k3_code") is not None and r[col_idx["k3_code"]] is not None else ""
        if not k3:
            stats["skipped"] += 1
            continue
        def _cell(field):
            i = col_idx.get(field)
            if i is None or i >= len(r) or r[i] is None:
                return ""
            return str(r[i]).strip()
        name = _cell("part_name")
        supplier = _cell("supplier")
        claim_req = _cell("claim_requirement")
        e10 = _cell("e10_code")
        erp_item = _cell("erp_item")
        if claim_req and claim_req not in VALID_CLAIM_REQUIREMENTS:
            stats["errors"].append(f"[{k3}] 索赔需求「{claim_req}」非法，已忽略该字段（合法：{'/'.join(VALID_CLAIM_REQUIREMENTS)}）")
            claim_req = ""
        if supplier and supplier not in supp_cache:
            sname = None
            existing = frappe.db.get_value("Supplier", {"supplier_name": supplier}, "name")
            if existing:
                sname = existing
            else:
                group = "All Supplier Groups"
                if not frappe.db.exists("Supplier Group", group):
                    frappe.get_doc({"doctype": "Supplier Group", "supplier_group_name": group}).insert(ignore_permissions=True)
                doc = frappe.get_doc({"doctype": "Supplier", "supplier_name": supplier, "supplier_group": group})
                doc.insert(ignore_permissions=True)
                sname = doc.name
                created_suppliers.append(supplier)
            supp_cache[supplier] = sname
        # 校验 ERP 物料（若提供）
        if erp_item and not frappe.db.exists("Item", erp_item):
            stats["errors"].append(f"[{k3}] ERP物料「{erp_item}」在 ERP 中不存在，erp_item 留空（可后续在主档补充）")
            erp_item = ""
        if not erp_item:
            erp_item = frappe.db.get_value("Spare Part", {"k3_code": k3}, "erp_item") or ""
        updates = {
            "e10_code": e10 or None,
            "part_name": name,
            "supplier": supp_cache.get(supplier, "") or None,
            "claim_requirement": claim_req or None,
            "erp_item": erp_item or None,
        }
        existing_doc = frappe.db.exists("Spare Part", {"k3_code": k3})
        if existing_doc:
            doc = frappe.get_doc("Spare Part", existing_doc)
            changed = False
            for f, v in updates.items():
                if v is not None and doc.get(f) != v:
                    doc.set(f, v)
                    changed = True
            if changed:
                doc.save(ignore_permissions=True)
                stats["updated"] += 1
            else:
                stats["skipped"] += 1
        else:
            data = {"doctype": "Spare Part", "k3_code": k3}
            data.update({f: v for f, v in updates.items() if v is not None})
            frappe.get_doc(data).insert(ignore_permissions=True)
            stats["created"] += 1
    frappe.db.commit()
    stats["suppliers_created"] = created_suppliers
    return stats
